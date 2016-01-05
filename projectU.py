#!/usr/bin/env python
import os
import re
import signal
import subprocess
import sys
from argparse import ArgumentParser

import openpyxl
from BeautifulSoup import BeautifulSoup

CASE_FOLDER = "%s/launchtime/" % sys.path[0]
d_realtask = {}
cur_task = ""
terminate_executor = False


class position:
    def __init__(self, x, y):
        self.y = str(y)
        self.x = x

    def right(self, y=None, step=1):
        self.x = chr(ord(self.x) + step)
        if y is not None:
            self.y = y
        return "%s%s" % (self.x, self.y)

    def left(self, step=1, y=None):
        self.x = chr(ord(self.x) - step)
        if y is not None:
            self.y = y
        return "%s%s" % (self.x, self.y)

    def down(self, step=1, x=None):
        if x is not None:
            self.x = x
        self.y = str(int(self.y) + step)
        return "%s%s" % (self.x, self.y)

    def up(self, step=1, x=None):
        if x is not None:
            self.x = x
        self.y = str(int(self.y) - step)
        return "%s%s" % (self.x, self.y)

    def move(self, x, y):
        self.y = str(y)
        self.x = x

    def cur(self):
        return "%s%s" % (self.x, self.y)


class alphachar:
    def __init__(self, s):
        self.s = s

    def __str__(self):
        return self.s

    def __add__(self, other):
        return chr(ord(self.s) + other)

    def __sub__(self, other):
        return chr(ord(self.s) - other)


def parse_tasklist(fname="tasklist.xml"):
    if os.path.exists(fname):
        with open(fname, "r") as f:
            soup = BeautifulSoup("\n".join(f.readlines()))
            l_node = soup.findAll('case')
            d_task = {}
            # add index to distinguish same cases in on tasklist
            index = 0
            for node in l_node:
                d_attr = {}
                for key, val in node.attrs:
                    d_attr[key] = val
                args = []
                args.append(d_attr['name'])
                args.append('-t %s' % d_attr['time'])
                args.append('-r %s' % d_attr['repeat'])
                if d_attr['warm_launch'] == 'true':
                    args.append('-w')
                if d_attr['systrace'] != '-1':
                    args.append('--systrace %s' % d_attr['systrace'])
                d_task["%s_%d" % (d_attr['name'], index)] = (" ".join(args))
                index += 1
            return d_task
    else:
        print "%s not found in current folder" % fname
        sys.exit(1)


def update_tasklist(d_task, fname="tasklist.xml"):
    soup = ""
    with open(fname, "r") as f:
        soup = BeautifulSoup("\n".join(f.readlines()))
        for node in soup.findAll('case'):
            if node['name'] not in d_task:
                print node['name']
                node.extract()
    with open(fname, 'w') as f:
        f.write("\n".join(map(str, soup.contents)))


def task_executor(d_task):
    global real_tasklist
    global cur_task
    d_realtask = dict(d_task)
    for key in d_task:
        cur_task = subprocess.Popen("python %s" % (d_task[key]), shell=True, preexec_fn=os.setsid)
        cur_task.wait()
        if cur_task.returncode == 0:
            del d_realtask[key]
            update_tasklist(d_realtask)
        if terminate_executor:
            break


def generate_xlsx():
    result_dict = {}
    for p, d, f in os.walk("%s/dmesg" % os.getcwd()):
        for t in f:
            group = re.match('(?P<case>[a-zA-Z_\(\)]+)-.*\((?P<result>\d+).*', t)
            if group is not None:
                if group.group("case") not in result_dict:
                    result_dict[group.group("case")] = []
                result_dict[group.group("case")].append(group.group("result"))

    wb = openpyxl.workbook.Workbook()
    ws_raw = wb.create_sheet(title='raw_data')
    ws_summary = wb.create_sheet(title='summary')
    pos = position('B', 1)
    max_result_count = 0
    for case in result_dict:
        ws_raw[pos.cur()] = case
        pos.down()
        case_results = result_dict[case]
        max_result_count = max(max_result_count, len(case_results))
        for result in case_results:
            ws_raw[pos.down()] = int(result)
        ws_raw[pos.up(step=len(case_results))] = '=MEDIAN(%s:%s)' % (pos.up(step=len(case_results) - 1), pos.down(step=len(case_results) - 1))
        pos.up()
        pos.right()
    pos.move('A', 2)
    ws_raw[pos.cur()] = 'MEDIAN'
    for i in xrange(max_result_count):
        ws_raw[pos.down()] = "Result%d" % (i + 1)

    quote_column = alphachar('B')
    pos.move('A', 1)
    case_count = len(result_dict.keys())
    for i in xrange(case_count):
        ws_summary[pos.cur()] = "=ws_raw!%s1" % (quote_column + i)
        ws_summary[pos.right()] = "=ws_raw!%s2" % (quote_column + i)
        pos.down(x='A')

    wb.save("tmp.xlsx")


def scan_cases():
    case_dict = {}
    for p, d, f in os.walk(CASE_FOLDER):
        for py in f:
            if py[-2:] == "py" and py != 'launch_script_generator.py':
                case_dict[py] = os.path.join(p, py)
    return case_dict


def generate_tasklist(case_dict):
    case_node = "<case name='%s' time='5' repeat='5' warm_launch='false' systrace='-1'/>"
    with open('tasklist.xml', 'w') as f:
        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes" ?><hierarchy rotation="0">')
        f.write('\n')
        for case in case_dict:
            f.write(case_node % case_dict[case])
            f.write('\n')
        f.write('</hierarchy>')


def main():
    p = ArgumentParser(usage='projectU.py [-u | -g]', description='Author wxl')
    p.add_argument('-u', action='store_true',  dest='update_only', help='only update the xlsx file')
    p.add_argument('-g', action='store_true', dest='generate_tasklist_only', help='only generate tasklist')
    a = p.parse_known_args(sys.argv)
    update_only = a[0].update_only
    generate_tasklist_only = a[0].generate_tasklist_only
    if update_only:
        generate_xlsx()
        return 0
    if generate_tasklist_only:
        generate_tasklist(scan_cases())
        return 0

    global terminate_executor
    d_task = parse_tasklist()
    try:
        task_executor(d_task)
    except KeyboardInterrupt:
        os.killpg(cur_task.pid, signal.SIGTERM)
        terminate_executor = True
    generate_xlsx()


if __name__ == '__main__':
    main()
